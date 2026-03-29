"""客户数据管理模块 - 存储、去重、导出Excel"""
import os
import json
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.table import Table

from config import DATA_DIR, OUTPUT_DIR, EXCEL_OUTPUT
from models import Lead

console = Console()


def ensure_dirs():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def save_leads_json(leads: list[Lead], filename: str = "leads.json"):
    """保存线索到JSON文件"""
    ensure_dirs()
    filepath = os.path.join(DATA_DIR, filename)

    existing = []
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            existing = json.load(f)

    new_data = [lead.to_dict() for lead in leads]
    existing.extend(new_data)

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    console.print(f"[green]已保存 {len(new_data)} 条线索到 {filepath}（总计 {len(existing)} 条）[/green]")


def load_leads_json(filename: str = "leads.json") -> list[Lead]:
    """从JSON加载线索"""
    filepath = os.path.join(DATA_DIR, filename)
    if not os.path.exists(filepath):
        return []

    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    return [Lead(**item) for item in data]


def deduplicate(leads: list[Lead]) -> list[Lead]:
    """按公司名去重，保留信息最完整的记录"""
    seen: dict[str, Lead] = {}
    for lead in leads:
        name = lead.company_name.strip()
        if not name:
            continue
        if name in seen:
            old = seen[name]
            if not old.phone and lead.phone:
                old.phone = lead.phone
            if not old.email and lead.email:
                old.email = lead.email
            if not old.address and lead.address:
                old.address = lead.address
            if not old.contact_person and lead.contact_person:
                old.contact_person = lead.contact_person
            if not old.website and lead.website:
                old.website = lead.website
        else:
            seen[name] = lead

    result = list(seen.values())
    console.print(f"[yellow]去重: {len(leads)} -> {len(result)} 条[/yellow]")
    return result


def export_excel(leads: list[Lead], filepath: str = None):
    """导出客户名单到格式化的Excel文件"""
    ensure_dirs()
    if filepath is None:
        filepath = EXCEL_OUTPUT

    wb = Workbook()
    ws = wb.active
    ws.title = "潜在客户名单"

    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = Lead.headers()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)

    for row_idx, lead in enumerate(leads, 2):
        for col_idx, value in enumerate(lead.to_row(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    col_widths = [25, 12, 10, 18, 25, 30, 18, 30, 40, 10, 20, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(filepath)
    console.print(f"[bold green]已导出 {len(leads)} 条客户数据到: {filepath}[/bold green]")

def show_summary(leads: list[Lead]):
    """在终端展示客户数据摘要"""
    if not leads:
        console.print("[yellow]暂无客户数据[/yellow]")
        return

    table = Table(title=f"客户名单摘要（共 {len(leads)} 条）", show_lines=True)
    table.add_column("#", style="dim", width=4)
    table.add_column("公司名称", style="bold", width=25)
    table.add_column("行业", width=10)
    table.add_column("联系人", width=8)
    table.add_column("电话", width=15)
    table.add_column("邮箱", width=22)
    table.add_column("来源", width=15)

    for i, lead in enumerate(leads[:50], 1):
        table.add_row(
            str(i),
            lead.company_name[:20],
            lead.industry,
            lead.contact_person or "-",
            lead.phone or "-",
            lead.email or "-",
            lead.source[:12],
        )

    console.print(table)

    has_phone = sum(1 for l in leads if l.phone)
    has_email = sum(1 for l in leads if l.email)
    console.print(f"\n[cyan]有电话: {has_phone}/{len(leads)} | 有邮箱: {has_email}/{len(leads)}[/cyan]")

    industry_counts: dict[str, int] = {}
    for l in leads:
        ind = l.industry or "未分类"
        industry_counts[ind] = industry_counts.get(ind, 0) + 1

    console.print("\n[bold]行业分布:[/bold]")
    for ind, count in sorted(industry_counts.items(), key=lambda x: -x[1]):
        console.print(f"  {ind}: {count} 家")
